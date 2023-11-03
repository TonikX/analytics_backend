import openpyxl
from openpyxl.styles import Alignment, PatternFill, Side, Border, Font
from sentry_sdk import capture_exception

from analytics_project.settings import AP_FILE_ROUTE
from gia_practice_app.GIA.models import GIA
from gia_practice_app.GIA.serializers import GIASmallSerializer
from gia_practice_app.Practice.models import Practice
from gia_practice_app.Practice.serializers import PracticeCompetenceSerializer
from workprogramsapp.disciplineblockmodules.ze_module_logic import recursion_module, generate_full_ze_list, \
    recursion_module_per_ze, sum_lists
from workprogramsapp.educational_program.general_prof_competencies.models import \
    GeneralProfCompetencesInGroupOfGeneralCharacteristic
from workprogramsapp.educational_program.general_prof_competencies.serializers import \
    GeneralProfCompetencesInGroupOfGeneralCharacteristicSerializer
from workprogramsapp.educational_program.key_competences.models import KeyCompetencesInGroupOfGeneralCharacteristic
from workprogramsapp.educational_program.key_competences.serializers import \
    KeyCompetencesInGroupOfGeneralCharacteristicSerializer
from workprogramsapp.educational_program.over_professional_competencies.models import \
    OverProfCompetencesInGroupOfGeneralCharacteristic
from workprogramsapp.educational_program.over_professional_competencies.serializers import \
    OverProfCompetencesInGroupOfGeneralCharacteristicSerializer
from workprogramsapp.educational_program.pk_comptencies.models import PkCompetencesInGroupOfGeneralCharacteristic
from workprogramsapp.educational_program.pk_comptencies.serializers import \
    PkCompetencesInGroupOfGeneralCharacteristicSerializer
from workprogramsapp.educational_program.process_modules_for_matrix import get_competences_wp, get_competences_practice
from workprogramsapp.educational_program.serializers import WorkProgramCompetenceIndicatorSerializer
from workprogramsapp.educational_program.views import GetCompetenceMatrix
from workprogramsapp.models import DisciplineBlock, DisciplineBlockModule, WorkProgramChangeInDisciplineBlockModule, \
    СertificationEvaluationTool, ImplementationAcademicPlan, FieldOfStudy, GeneralCharacteristics, WorkProgram, \
    WorkProgramInFieldOfStudy, PracticeInFieldOfStudy
from workprogramsapp.serializers import AcademicPlanSerializer

color_list = ["4c69a9", "5f84d4", "6f90d8", "8fa8e0", "afc1e9", "cfdaf2", "dfe6f6", "eff2fa", "eff2fa", "eff2fa",
              "eff2fa"]


def fill_row(ws, level, color, bold=False):
    for i in range(1, 65):
        cell = ws.cell(row=level, column=i)
        cell.fill = PatternFill(start_color=color, fill_type="solid")
        side = Side(border_style='thin', color='000000')
        cell.border = Border(top=side, bottom=side, left=side, right=side)
        font = Font(name='Times New Roman', size=12, bold=bold, color='000000')
        cell.font = font


def insert_cell_data(ws, level, column, data, horizontal='center'):
    cell = ws.cell(row=level, column=column)
    cell.value = data
    cell.alignment = Alignment(wrap_text=True, horizontal=horizontal, vertical='center')


def insert_cell_data_range(ws, level, start, finish, data_list):
    i = 0
    for col in range(start, finish + 1):
        cell = ws.cell(row=level, column=col)
        try:
            cell.value = data_list[i]
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            i += 1
        except IndexError:
            pass


def find_competence_row(competence_dict, competences_in_program):
    plus_index_list = []
    for comp in competences_in_program["competences"]["competences"]:
        for key in competence_dict:
            for i, comp_in_header in enumerate(competence_dict[key]["queryset"]):
                if comp["id"] == comp_in_header.competence.id:
                    plus_index_list.append(competence_dict[key]["start"] + i)
    return plus_index_list


def process_changeblock(ws, level, block_module, unique_dict, competence_dict):
    for change_block in WorkProgramChangeInDisciplineBlockModule.objects.filter(
            discipline_block_module=block_module):
        for work_program in WorkProgram.objects.filter(work_program_in_change_block=change_block):
            if (work_program.id not in unique_dict["unique_wp"]) or (unique_dict["is_first_iter_ap"]):
                wp_in_fs = WorkProgramInFieldOfStudy.objects.get(work_program=work_program,
                                                                 work_program_change_in_discipline_block_module=change_block)

                competences_in_program = {"id": work_program.id, "title": work_program.title,
                                          "competences": get_competences_wp(wp_in_fs)}
                unique_dict["unique_wp"].append(work_program.id)
                fill_row(ws, level, "FFFFFF")
                insert_cell_data(ws=ws, level=level, column=1, data=work_program.title, horizontal="left")
                for plus_column in find_competence_row(competence_dict, competences_in_program):
                    insert_cell_data(ws=ws, level=level, column=plus_column, data="+", horizontal="left")
                level += 1

            else:
                pass
                # print(work_program)
        for practice in Practice.objects.filter(practice_in_change_block=change_block):
            if (practice.id not in unique_dict["unique_practice"]) or (unique_dict["is_first_iter_ap"]):
                practice_in_fs = PracticeInFieldOfStudy.objects.get(practice=practice,
                                                                    work_program_change_in_discipline_block_module=change_block)
                competences_in_program = {"id": practice.id, "title": practice.title,
                                          "competences": get_competences_practice(practice_in_fs)}
                unique_dict["unique_practice"].append(practice.id)
                fill_row(ws, level, "FFFFFF")
                insert_cell_data(ws=ws, level=level, column=1, data=practice.title, horizontal="left")
                for plus_column in find_competence_row(competence_dict, competences_in_program):
                    insert_cell_data(ws=ws, level=level, column=plus_column, data="+", horizontal="left")
                level += 1
            else:
                pass
                # print(work_program)
        for gia in GIA.objects.filter(gia_in_change_block=change_block):
            if (gia.id not in unique_dict["unique_gia"]) or (unique_dict["is_first_iter_ap"]):
                serializer = GIASmallSerializer(gia)
                unique_dict["unique_gia"].append(gia.id)
                fill_row(ws, level, "FFFFFF")
                insert_cell_data(ws=ws, level=level, column=1, data=gia.title, horizontal="left")
                level += 1
            else:
                pass
    return level


def recursion_module_matrix(ws, level, modules, unique_dict, competence_dict, depth):
    # print(level)
    for block_module in modules:
        fill_row(ws, level, color_list[depth])
        insert_cell_data(ws=ws, level=level, column=1, data=block_module.name, horizontal="left")
        level += 1

        childs = block_module.childs.all()
        if childs.exists():
            level = recursion_module_matrix(ws, level, childs, unique_dict, competence_dict, depth + 1)
        else:
            level = process_changeblock(ws, level, block_module, unique_dict, competence_dict)

    return level


def process_excel_competence_matrix(gen_characteristic):
    #wb_obj = openpyxl.load_workbook("C:\\Users\\s4\\Desktop\\analytics_backend\\application\\static-backend\\export_template\\competence_matrix_template_2023.xlsx")
    wb_obj = openpyxl.load_workbook("/application/static-backend/export_template/competence_matrix_template_2023.xlsx")

    unique_dict = \
        {"is_first_iter_ap": True,
         "unique_wp": [],
         "unique_practice": [],
         "unique_gia": []}

    ws = wb_obj["МК"]
    level = 3
    depth = 0

    competence_dict = competence_header(ws, gen_characteristic)

    academic_plans = gen_characteristic.educational_program.all()
    for ap in academic_plans:
        academic_plan = ap.academic_plan
        for block in DisciplineBlock.objects.filter(academic_plan=academic_plan).order_by("name"):
            insert_cell_data(ws=ws, level=level, column=1, data=block.name, horizontal="left")
            fill_row(ws, level, "394f7f")

            level += 1

            modules = DisciplineBlockModule.objects.filter(descipline_block=block)
            level = recursion_module_matrix(ws, level, modules, unique_dict, competence_dict, depth + 1)
        unique_dict["is_first_iter_ap"] = False

    return wb_obj
    # return wb_obj


def fill_competences(ws, column_start, competence_queryset, competence_dict, name_comp, string_comp):
    level = 2

    # Генерация номеров для профессионапльных компетенций
    if not name_comp == "pk_competences":
        competences_list = [comp.competence.number for comp in competence_queryset]
    else:
        pk_dict = {} #Словарь для поиска одинаковыз компетенций
        competences_list = []
        pk_code_inc = 0
        for pk_type in ["prof", "fore", "min"]:
            for comp in competence_queryset.filter(group_of_pk__type_of_pk_competence=pk_type):
                if not pk_dict.get(comp.competence.id):
                    pk_code_inc += 1
                    pk_dict[comp.competence.id] = pk_code_inc
                    temp_pk_code = pk_code_inc
                else:
                    temp_pk_code = pk_dict[comp.competence.id]
                competences_list.append(f"ПК-{temp_pk_code}")

    # Формируем данные
    competence_dict[name_comp]["start"] = column_start
    competence_dict[name_comp]["list_num"] = competences_list
    column_end = column_start + len(competences_list) - 1
    # Заполняем строки
    insert_cell_data_range(ws, level, column_start, column_end, competences_list)
    ws.merge_cells(start_row=1, end_row=1, start_column=column_start, end_column=column_end)
    insert_cell_data(ws=ws, level=1, column=column_start, data=string_comp)
    # Ставим границу
    cell = ws.cell(row=1, column=column_end)

    side = Side(border_style='thick', color='000000')
    cell.border = Border(right=side)

    cell = ws.cell(row=2, column=column_end)
    cell.border = Border(right=side)

    return column_end + 1


def competence_header(ws, gen_characteristic):
    #gen_characteristic = GeneralCharacteristics.objects.get(pk=gen_pk)
    """pk_competences = PkCompetencesInGroupOfGeneralCharacteristic.objects.filter(
        group_of_pk__general_characteristic=gen_characteristic, competence__isnull=False).distinct()"""
    pk_competences = PkCompetencesInGroupOfGeneralCharacteristic.objects.filter(
        group_of_pk__general_characteristic=gen_characteristic, competence__isnull=False,
        group_of_pk__type_of_pk_competence="prof").distinct() | PkCompetencesInGroupOfGeneralCharacteristic.objects.filter(
        group_of_pk__general_characteristic=gen_characteristic, competence__isnull=False,
        group_of_pk__type_of_pk_competence="fore").distinct() | PkCompetencesInGroupOfGeneralCharacteristic.objects.filter(
        group_of_pk__general_characteristic=gen_characteristic, competence__isnull=False,
        group_of_pk__type_of_pk_competence="min").distinct()
    general_prof_competences = GeneralProfCompetencesInGroupOfGeneralCharacteristic.objects.filter(
        group_of_pk__educational_standard=gen_characteristic.educational_standard,
        competence__isnull=False)
    key_competences = KeyCompetencesInGroupOfGeneralCharacteristic.objects.filter(
        group_of_pk__educational_standard=gen_characteristic.educational_standard,
        competence__isnull=False).distinct()
    over_prof_competences = OverProfCompetencesInGroupOfGeneralCharacteristic.objects.filter(
        group_of_pk__educational_standard=gen_characteristic.educational_standard,
        competence__isnull=False).distinct()
    competence_dict = {
        "key_competences": {"start": 0, "list_num": [], "queryset": key_competences,
                            "label": "Ключевые компетенции"},
        "over_prof_competences": {"start": 0, "list_num": [], "queryset": over_prof_competences,
                                  "label": "Надпрофессиональные компетенции"},
        "general_prof_competences": {"start": 0, "list_num": [], "queryset": general_prof_competences,
                                     "label": "Общепрофессиональные компетенции "},
        "pk_competences": {"start": 0, "list_num": [], "queryset": pk_competences,
                           "label": "Профессиональные компетенции"}
    }
    column_start = 2
    for key in competence_dict:
        column_start = fill_competences(ws=ws, column_start=column_start,
                                        competence_queryset=competence_dict[key]["queryset"],
                                        competence_dict=competence_dict,
                                        name_comp=key,
                                        string_comp=competence_dict[key]["label"])

    cell = ws.cell(row=1, column=2)
    side = Side(border_style='thick', color='000000')
    cell.border = Border(left=side)

    cell = ws.cell(row=2, column=2)
    cell.border = Border(left=side)

    return competence_dict
