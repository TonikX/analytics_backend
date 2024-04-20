import openpyxl
from openpyxl.styles import Alignment, PatternFill, Side, Border, Font
from sentry_sdk import capture_exception

from analytics_project.settings import AP_FILE_ROUTE
from workprogramsapp.disciplineblockmodules.ze_module_logic import (
    generate_full_ze_list,
    recursion_module,
    recursion_module_per_ze,
    sum_lists,
)
from workprogramsapp.models import (
    CertificationEvaluationTool,
    DisciplineBlock,
    DisciplineBlockModule,
    FieldOfStudy,
    ImplementationAcademicPlan,
    WorkProgramChangeInDisciplineBlockModule,
)

color_list = [
    "4c69a9",
    "5f84d4",
    "6f90d8",
    "8fa8e0",
    "afc1e9",
    "cfdaf2",
    "dfe6f6",
    "eff2fa",
    "eff2fa",
    "eff2fa",
    "eff2fa",
]

columns_dict = {
    "wp_id": {"column": 1},
    "choosing_rule": {"column": 2},
    "choosing_param": {"column": 3},
    "choosing_term": {"column": 4},
    "choosing_format": {"column": 5},
    "replaceable": {"column": 6},
    "name": {"column": 7},
    "ze_all": {"column": 8},
    "hours_all": {"column": 9},
    "ze_by_term": {"column_start": 10, "column_finish": 17},
    "exam": {"column": 18},
    "diff": {"column": 19},
    "credit": {"column": 20},
    "course_project": {"column": 21},
    "course_work": {"column": 22},
    "contact_work": {"column": 23},
    "seminary_sum": {"column": 24},
    "lecture": {"column": 25},
    "labs": {"column": 26},
    "practice": {"column": 27},
    "consultations": {"column": 28},
    "srs": {"column": 29},
    "hours_by_term": {"column": 30},
    "realizer": {"column": 62},
    "percent_seminary": {"column": 63},
    "realisation_language": {"column": 64},
}
ERR_DICT = {"wp_err": []}


def fill_row(ws, level, color, bold=False):
    for i in range(1, 65):
        cell = ws.cell(row=level, column=i)
        cell.fill = PatternFill(start_color=color, fill_type="solid")
        side = Side(border_style="thin", color="000000")
        cell.border = Border(top=side, bottom=side, left=side, right=side)
        font = Font(name="Times New Roman", size=12, bold=bold, color="000000")
        cell.font = font


def insert_cell_data(
    ws, level, column_name, data, do_line_merge=False, horizontal="center"
):
    cell = ws.cell(row=level, column=columns_dict[column_name]["column"])
    cell.value = data
    cell.alignment = Alignment(wrap_text=True, horizontal=horizontal, vertical="center")
    if do_line_merge:
        ws.merge_cells(
            start_row=level,
            start_column=columns_dict[column_name]["column"],
            end_row=level + 1,
            end_column=columns_dict[column_name]["column"],
        )


def insert_cell_data_range(ws, level, column_name, data_list):
    i = 0
    for col in range(
        columns_dict[column_name]["column_start"],
        columns_dict[column_name]["column_finish"] + 1,
    ):
        cell = ws.cell(row=level, column=col)
        try:
            cell.value = data_list[i]
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )
            i += 1
        except IndexError:
            pass


def process_evaluation_tools(ws, level, wp=None, module=None, start_sem=0):
    if wp:
        tools = CertificationEvaluationTool.objects.filter(work_program=wp)
    elif module:
        tools = CertificationEvaluationTool.objects.filter(
            discipline_block_module=module
        )
    else:
        tools = CertificationEvaluationTool.objects.none()
    try:
        exam_tools = "".join(
            map(str, [tool.semester + start_sem for tool in tools.filter(type="1")])
        )
        diff_tools = "".join(
            map(str, [tool.semester + start_sem for tool in tools.filter(type="2")])
        )
        credit_tools = "".join(
            map(str, [tool.semester + start_sem for tool in tools.filter(type="3")])
        )
        course_work_tools = "".join(
            map(str, [tool.semester + start_sem for tool in tools.filter(type="4")])
        )
        course_project_tools = "".join(
            map(str, [tool.semester + start_sem for tool in tools.filter(type="5")])
        )
        insert_cell_data(ws, level, "exam", exam_tools)
        insert_cell_data(ws, level, "diff", diff_tools)
        insert_cell_data(ws, level, "credit", credit_tools)
        insert_cell_data(ws, level, "course_project", course_project_tools)
        insert_cell_data(ws, level, "course_work", course_work_tools)

    except Exception as e:
        capture_exception(e)


def process_hours_by_terms(ws, level, hours_list, offset=0):
    start_hours = columns_dict["hours_by_term"]["column"] + offset
    for i, hour in enumerate(hours_list):
        if hour != 0:
            cell = ws.cell(row=level, column=start_hours + i * 4)
            cell.value = hours_list[i]
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )


def process_gia(changeblock, level, ws):
    for gia in changeblock.gia.all():
        fill_row(ws, level, "FFFFFF")
        insert_cell_data(
            ws=ws, level=level, column_name="wp_id", data=gia.discipline_code
        )
        try:
            insert_cell_data(
                ws=ws,
                level=level,
                column_name="choosing_term",
                data=", ".join(map(str, changeblock.semester_start)),
            )

        except Exception as e:
            capture_exception(e)
        insert_cell_data(
            ws=ws, level=level, column_name="name", data=gia.title, horizontal="left"
        )

        try:
            ze = [int(unit) for unit in gia.ze_v_sem.split(", ")]
            sum_ze = sum(ze)
            insert_cell_data(ws=ws, level=level, column_name="ze_all", data=sum_ze)
            insert_cell_data(
                ws=ws, level=level, column_name="hours_all", data=sum_ze * 36
            )
            wp_ze_by_term = generate_full_ze_list(ze, changeblock.semester_start)[0]
            insert_cell_data_range(ws, level, "ze_by_term", wp_ze_by_term)
        except Exception as e:
            capture_exception(e)

        insert_cell_data(ws, level, "exam", changeblock.semester_start[0])
        insert_cell_data(
            ws=ws,
            level=level,
            column_name="realizer",
            data=gia.structural_unit.short_name,
        )
        level += 1
        return level


def process_practice(changeblock, level, ws):
    for practice in changeblock.practice.all():
        tool_types = {1: "", 2: "", 3: "", 4: "", 5: ""}
        fill_row(ws, level, "FFFFFF")
        insert_cell_data(
            ws=ws, level=level, column_name="wp_id", data=practice.discipline_code
        )
        try:
            insert_cell_data(
                ws=ws,
                level=level,
                column_name="choosing_term",
                data=", ".join(map(str, changeblock.semester_start)),
            )

        except Exception as e:
            capture_exception(e)
        insert_cell_data(
            ws=ws,
            level=level,
            column_name="name",
            data=practice.title,
            horizontal="left",
        )

        try:
            ze = [int(unit) for unit in practice.ze_v_sem.split(", ")]
            sum_ze = sum(ze)
            insert_cell_data(ws=ws, level=level, column_name="ze_all", data=sum_ze)
            insert_cell_data(
                ws=ws, level=level, column_name="hours_all", data=sum_ze * 36
            )
            wp_ze_by_term = generate_full_ze_list(ze, changeblock.semester_start)[0]
            insert_cell_data_range(ws, level, "ze_by_term", wp_ze_by_term)
        except Exception as e:
            capture_exception(e)
        try:
            list_of_tools = eval(practice.evaluation_tools_v_sem)

            for i, el in enumerate(list_of_tools):
                for tool_type in el:
                    tool_types[tool_type] += str(i + changeblock.semester_start[0])
            insert_cell_data(ws, level, "exam", tool_types[1])
            insert_cell_data(ws, level, "diff", tool_types[2])
            insert_cell_data(ws, level, "credit", tool_types[3])
            insert_cell_data(ws, level, "course_project", tool_types[4])
            insert_cell_data(ws, level, "course_work", tool_types[5])
        except Exception as e:
            capture_exception(e)

        insert_cell_data(
            ws=ws,
            level=level,
            column_name="realizer",
            data=practice.structural_unit.short_name,
        )
        if practice.language:
            insert_cell_data(
                ws=ws,
                level=level,
                column_name="realisation_language",
                data=practice.language.upper(),
            )
        level += 1
        return level


def process_changeblock(changeblocks, level, ws):
    for changeblock in changeblocks:
        if changeblock.gia.all().exists():
            level = process_gia(changeblock, level, ws)
        if changeblock.practice.all().exists():
            level = process_practice(changeblock, level, ws)

        for wp in changeblock.work_program.all():
            fill_row(ws, level, "FFFFFF")

            insert_cell_data(
                ws=ws, level=level, column_name="wp_id", data=wp.discipline_code
            )
            insert_cell_data(
                ws=ws,
                level=level,
                column_name="choosing_term",
                data=", ".join(map(str, changeblock.semester_start)),
            )
            if wp.implementation_format == "offline":
                choosing_format = "оф"
            elif wp.implementation_format == "online":
                choosing_format = "он"
            elif wp.implementation_format == "mixed":
                choosing_format = "микс"
            else:
                choosing_format = "оф"
            insert_cell_data(
                ws=ws, level=level, column_name="choosing_format", data=choosing_format
            )
            replaceable = "+" if changeblock.change_type == "Optionally" else "-"
            insert_cell_data(
                ws=ws, level=level, column_name="replaceable", data=replaceable
            )

            insert_cell_data(
                ws=ws, level=level, column_name="name", data=wp.title, horizontal="left"
            )
            try:
                ze = [int(unit) for unit in wp.ze_v_sem.split(", ")]
                sum_ze = sum(ze)
                insert_cell_data(ws=ws, level=level, column_name="ze_all", data=sum_ze)
                insert_cell_data(
                    ws=ws, level=level, column_name="hours_all", data=sum_ze * 36
                )

                wp_ze_by_term = generate_full_ze_list(ze, changeblock.semester_start)[0]
                insert_cell_data_range(ws, level, "ze_by_term", wp_ze_by_term)
            except AttributeError:
                ERR_DICT["wp_err"].append(
                    f'В РПД {wp.id} "{wp.title}" зачетные единицы указаны в некорректном формате'
                )
            except Exception as e:
                capture_exception(e)
            try:
                process_evaluation_tools(
                    ws, level, wp, None, changeblock.semester_start[0] - 1
                )
            except IndexError as e:
                ERR_DICT["wp_err"].append(
                    f'В РПД {wp.id} "{wp.title}" не указан семестр начала изучения дисциплины'
                )

            try:
                extend_list = [0 for _ in range(10)]
                lecture_list = [float(hour) for hour in wp.lecture_hours_v2.split(", ")]
                practice_list = [
                    float(hour) for hour in wp.practice_hours_v2.split(", ")
                ]
                sro_list = [float(hour) for hour in wp.srs_hours_v2.split(", ")]
                lab_list = [float(hour) for hour in wp.lab_hours_v2.split(", ")]
                try:
                    cons_list = [float(hour) for hour in wp.consultation_v2.split(", ")]
                    insert_cell_data(
                        ws=ws,
                        level=level,
                        column_name="consultations",
                        data=sum(cons_list),
                    )
                except Exception as e:
                    cons_list = [0]
                    """
                    ERR_DICT["wp_err"].append(f"В РПД {wp.id} часы консультаций отсутсвуют"
                                              f" или указаны в неверном формате")
                    """

                contact_hours = round(
                    (
                        sum(lecture_list)
                        + sum(practice_list)
                        + sum(lab_list)
                        + sum(cons_list)
                    )
                    * 1.1,
                    2,
                )
                insert_cell_data(
                    ws=ws, level=level, column_name="contact_work", data=contact_hours
                )

                seminary_hours = (
                    sum(lecture_list)
                    + +sum(practice_list)
                    + sum(lab_list)
                    + sum(cons_list)
                )
                insert_cell_data(
                    ws=ws, level=level, column_name="seminary_sum", data=seminary_hours
                )

                insert_cell_data(
                    ws=ws, level=level, column_name="lecture", data=sum(lecture_list)
                )
                insert_cell_data(
                    ws=ws, level=level, column_name="labs", data=sum(lab_list)
                )
                insert_cell_data(
                    ws=ws, level=level, column_name="practice", data=sum(practice_list)
                )
                insert_cell_data(
                    ws=ws, level=level, column_name="srs", data=sum(sro_list)
                )

                process_hours_by_terms(
                    ws,
                    level,
                    generate_full_ze_list(lecture_list, changeblock.semester_start)[0],
                    offset=0,
                )
                process_hours_by_terms(
                    ws,
                    level,
                    generate_full_ze_list(lab_list, changeblock.semester_start)[0],
                    offset=1,
                )
                process_hours_by_terms(
                    ws,
                    level,
                    generate_full_ze_list(practice_list, changeblock.semester_start)[0],
                    offset=2,
                )
                process_hours_by_terms(
                    ws,
                    level,
                    generate_full_ze_list(cons_list, changeblock.semester_start)[0],
                    offset=3,
                )
            except Exception as e:
                capture_exception(e)

            insert_cell_data(
                ws=ws,
                level=level,
                column_name="realizer",
                data=wp.structural_unit.short_name,
            )

            try:
                insert_cell_data(
                    ws=ws,
                    level=level,
                    column_name="percent_seminary",
                    data=round(float((seminary_hours) / (sum_ze * 36)) * 100, 2),
                )
            except ZeroDivisionError:
                ERR_DICT["wp_err"].append(
                    f'В РПД {wp.id} "{wp.title}" не удалось подсчитать сумму зачетных единиц. '
                    f"Проверьте, указаны ли они в рабочей программе корректно"
                )
            except Exception as e:
                capture_exception(e)

            if wp.language:
                insert_cell_data(
                    ws=ws,
                    level=level,
                    column_name="realisation_language",
                    data=wp.language.upper(),
                )
            level += 1
    return level


def module_inside_recursion(modules, level, ws, depth=0):
    sum_data = {
        "level": level,
        "ze": 0,
        "ze_by_term": [0 for _ in range(10)],
        "contact_work": 0,
        "all_seminary": 0,
        "lecture": 0,
        "lab": 0,
        "practice": 0,
        "cons": 0,
        "srs": 0,
        "lecture_by_sem": [0 for _ in range(10)],
        "lab_by_sem": [0 for _ in range(10)],
        "practice_by_sem": [0 for _ in range(10)],
        "cons_by_sem": [0 for _ in range(10)],
    }

    for module in modules:

        fill_row(ws, level, color_list[depth])
        insert_cell_data(
            ws=ws,
            level=level,
            column_name="name",
            data=module.name,
            do_line_merge=False,
            horizontal="left",
        )

        if module.selection_rule == "choose_n_from_m":
            choosing_rule = "кол-во"
        elif module.selection_rule == "all":
            choosing_rule = "все"
        elif module.selection_rule == "any_quantity":
            choosing_rule = "все/0"
        else:
            choosing_rule = "з.е."

        insert_cell_data(
            ws=ws,
            level=level,
            column_name="choosing_rule",
            data=choosing_rule,
            do_line_merge=False,
        )
        insert_cell_data(
            ws=ws,
            level=level,
            column_name="choosing_param",
            data=module.selection_parametr,
            do_line_merge=False,
        )

        ze_module = recursion_module(module)
        insert_cell_data(
            ws=ws,
            level=level,
            column_name="ze_all",
            data=ze_module,
            do_line_merge=False,
        )
        insert_cell_data(
            ws=ws,
            level=level,
            column_name="hours_all",
            data=ze_module * 36,
            do_line_merge=False,
        )
        max_ze, max_hours_lab, max_hours_lec, max_hours_practice, max_hours_cons = (
            recursion_module_per_ze(module)
        )
        insert_cell_data_range(ws, level, "ze_by_term", max_ze)

        process_evaluation_tools(ws, level, module=module)

        # insert_cell_data_range(ws, level + 1, "ze_by_term", min_ze)
        process_hours_by_terms(ws, level, max_hours_lec, offset=0)
        process_hours_by_terms(ws, level, max_hours_lab, offset=1)
        process_hours_by_terms(ws, level, max_hours_practice, offset=2)

        level += 1
        if module.childs.all().exists():
            # depth += 1
            level = module_inside_recursion(module.childs.all(), level, ws, depth + 1)
        else:
            level = process_changeblock(
                WorkProgramChangeInDisciplineBlockModule.objects.filter(
                    discipline_block_module=module
                ),
                level,
                ws,
            )
        if depth == 0:
            sum_data["level"] = level
            sum_data["ze"] += ze_module
            sum_data["ze_by_term"] = sum_lists(sum_data["ze_by_term"], max_ze)
            # sum_data["contact_work"] += ze_module
            sum_data["lecture_by_sem"] = sum_lists(
                sum_data["lecture_by_sem"], max_hours_lec
            )
            sum_data["lab_by_sem"] = sum_lists(sum_data["lab_by_sem"], max_hours_lab)
            sum_data["practice_by_sem"] = sum_lists(
                sum_data["practice_by_sem"], max_hours_practice
            )
            # sum_data["cons_by_sem"] = sum_lists(sum_data["cons_by_sem"], max_hours_cons)
    if depth == 0:
        return sum_data
    else:
        return level


def process_excel(academic_plan):
    ERR_DICT["wp_err"] = []
    # wb_obj = openpyxl.load_workbook("C:\\Users\\123\\Desktop\\analitycs\\analytics_backend\\application\\workprogramsapp\\files_export\\plan.xlsx")
    wb_obj = openpyxl.load_workbook(AP_FILE_ROUTE)
    ws = wb_obj["УП"]
    start_list = 7
    final_ze_by_term = [0 for _ in range(10)]
    final_ze = 0
    for block in DisciplineBlock.objects.filter(academic_plan=academic_plan).order_by(
        "name"
    ):
        insert_cell_data(
            ws=ws,
            level=start_list,
            column_name="name",
            data=block.name,
            horizontal="left",
        )
        fill_row(ws, start_list, "394f7f")
        start_list += 1
        sum_data = module_inside_recursion(
            DisciplineBlockModule.objects.filter(descipline_block=block), start_list, ws
        )
        start_list = sum_data["level"]

        fill_row(ws, start_list, "899499")
        process_hours_by_terms(ws, start_list, sum_data["lecture_by_sem"], offset=0)
        process_hours_by_terms(ws, start_list, sum_data["lab_by_sem"], offset=1)
        process_hours_by_terms(ws, start_list, sum_data["practice_by_sem"], offset=2)
        # process_hours_by_terms(ws, start_list, sum_data["cons_by_sem"], offset=3)

        insert_cell_data(
            ws=ws,
            level=start_list,
            column_name="ze_all",
            data=sum_data["ze"],
            do_line_merge=False,
        )
        insert_cell_data(
            ws=ws,
            level=start_list,
            column_name="hours_all",
            data=sum_data["ze"] * 36,
            do_line_merge=False,
        )
        insert_cell_data_range(ws, start_list, "ze_by_term", sum_data["ze_by_term"])
        start_list += 1
        if block.name != "Блок 4. Факультативные модули (дисциплины)":
            final_ze += sum_data["ze"]
            final_ze_by_term = sum_lists(final_ze_by_term, sum_data["ze_by_term"])

    # imp = ImplementationAcademicPlan.objects.get(academic_plan=academic_plan)
    insert_cell_data(
        ws=ws, level=start_list, column_name="name", data="Объем ОП", horizontal="left"
    )
    fill_row(ws, start_list, "AB6E9D", bold=True)
    insert_cell_data(
        ws=ws,
        level=start_list,
        column_name="ze_all",
        data=final_ze,
        do_line_merge=False,
    )
    insert_cell_data(
        ws=ws,
        level=start_list,
        column_name="hours_all",
        data=final_ze * 36,
        do_line_merge=False,
    )
    insert_cell_data_range(ws, start_list, "ze_by_term", final_ze_by_term)

    ws = wb_obj["Титульный лист"]
    imp = ImplementationAcademicPlan.objects.filter(academic_plan=academic_plan)[0]
    fos = FieldOfStudy.objects.filter(
        implementation_academic_plan_in_field_of_study=imp
    )[0]
    insert_cell_data(
        ws=ws,
        level=4,
        column_name="wp_id",
        data=f'Наименование образовательной программы "{imp.title}"',
        do_line_merge=False,
    )
    insert_cell_data(
        ws=ws,
        level=5,
        column_name="wp_id",
        data=f"Направление подготовки: {fos.number} {fos.title}",
        do_line_merge=False,
    )
    insert_cell_data(
        ws=ws,
        level=8,
        column_name="wp_id",
        data=f"Очная форма обучения, срок получения образования - {imp.training_period} года, "
        f"год начала подготовки -  {imp.year}",
        do_line_merge=False,
    )
    if imp.language == "ru":
        language = "Русский"
    elif imp.language == "en":
        language = "Английский"
    elif imp.language == "kz":
        language = "Казахский"
    elif imp.language == "de":
        language = "Немецкий"
    else:
        language = "Русский/Английский"
    insert_cell_data(
        ws=ws,
        level=10,
        column_name="wp_id",
        data=f"Язык реализации ОП: {language}",
        do_line_merge=False,
    )

    return wb_obj, ERR_DICT
