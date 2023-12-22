import openpyxl
from openpyxl.styles import Alignment, PatternFill, Side, Border, Font
from sentry_sdk import capture_exception

from analytics_project.settings import AP_FILE_ROUTE
from gia_practice_app.Practice.models import Practice
from workprogramsapp.disciplineblockmodules.ze_module_logic import recursion_module, generate_full_ze_list, \
    recursion_module_per_ze, sum_lists
from workprogramsapp.models import DisciplineBlock, DisciplineBlockModule, WorkProgramChangeInDisciplineBlockModule, \
    СertificationEvaluationTool, ImplementationAcademicPlan, FieldOfStudy
from workprogramsapp.serializers import AcademicPlanSerializer


def fill_calendar(ws, sems_of_practice, type_practice, weeks=None, length=None):
    semester_grid = {1: {"start": [12, 3], "end": [12, 19]},
                     2: {"start": [12, 25], "end": [12, 42]},
                     3: {"start": [19, 3], "end": [19, 19]},
                     4: {"start": [19, 25], "end": [20, 42]},
                     5: {"start": [26, 3], "end": [26, 19]},
                     6: {"start": [26, 25], "end": [30, 41]},
                     7: {"start": [33, 3], "end": [35, 19]},
                     8: {"start": [33, 25], "end": [37, 37]},
                     }
    holidays_grid = [[13, 12], [12, 28], [16, 37], [17, 38],
                     [20, 12], [19, 28], [23, 37], [24, 38],
                     [27, 12], [26, 28], [30, 37], [31, 38],
                     [34, 12], [33, 28], [37, 37], [38, 38]]
    for i, sem in enumerate(sems_of_practice):
        grid = semester_grid[sem]
        if length:
            if length[i] == 0:
                continue
            column = grid["start"][1] + weeks[i] - 1
            column_end = grid["start"][1] + length[i] + weeks[i]
            row_end = grid["start"][0]
        else:
            column = grid["start"][1]
            row_end, column_end = grid["end"]

        while True:
            row = grid["start"][0]
            for i in range(1, 7):
                if row == row_end and column == column_end:
                    break
                cell = ws.cell(row=row, column=column)
                # print(cell.fill.start_color.index)
                if [row, column] not in holidays_grid:
                    if type_practice == "educational":
                        cell.fill = PatternFill(start_color="d6dce4", fill_type="solid")
                        if cell.fill.start_color.index == "99ccff":
                            cell.value = "п,уп"
                    elif type_practice == "production":
                        cell.fill = PatternFill(start_color="99ccff", fill_type="solid")
                        if cell.fill.start_color.index == "d6dce4":
                            cell.value = "п,уп"
                row += 1
            if row == row_end and column == column_end:
                break
            column += 1


"""
from workprogramsapp.files_export.calendar_export import *
from workprogramsapp.models import *
ap = AcademicPlan.objects.get(id=7302)
process_excel(ap)

"""


def process_practice(ws, practice_cb):
    try:
        practice_obj = Practice.objects.get(
            zuns_for_pr__work_program_change_in_discipline_block_module=practice_cb)
    except Practice.DoesNotExist:
        return
    if "преддипломная" in practice_obj.title.lower():
        return
    semester_start = practice_cb.semester_start[0]
    number_of_semesters = practice_obj.number_of_semesters
    semester_list = [i for i in range(semester_start, semester_start + number_of_semesters)]
    if practice_obj.format_practice == "dispersed":
        fill_calendar(ws, semester_list, practice_obj.kind_of_practice)
    elif practice_obj.format_practice == "dedicated":
        weeks_start = [int(week) for week in practice_obj.weeks_start.split(", ")]
        if not weeks_start:
            weeks_start = [1 for _ in range(number_of_semesters)]

        length = [float(ze) / 1.5 for ze in practice_obj.ze_v_sem.split(", ")]
        print(weeks_start, length)
        fill_calendar(ws, semester_list, practice_obj.kind_of_practice, weeks_start, length)


def process_excel_calendar(academic_plan):
    wb_obj = openpyxl.load_workbook(
        "C:\\Users\\s4\\Desktop\\analytics_backend\\application\\static-backend\\export_template\\calendar_2023.xlsx")
    # wb_obj = openpyxl.load_workbook(AP_FILE_ROUTE)
    ws = wb_obj["КУГ"]

    practice_cbs = academic_plan.get_all_changeblocks_from_ap(block_to_filter="Блок 2. Практика")
    modules_practice = DisciplineBlockModule.objects.filter(change_blocks_of_work_programs_in_modules__in=practice_cbs)
    for module in modules_practice:
        cbs = WorkProgramChangeInDisciplineBlockModule.objects.filter(discipline_block_module=module)
        if module.selection_rule == "choose_n_from_m":
            for i in range(int(module.selection_parametr)):
                practice_cb = cbs[i]
                process_practice(ws, practice_cb)
        if module.selection_rule == "by_credit_units":
            credit_units_counter = 0
            for practice_cb in cbs:
                try:
                    practice_obj = Practice.objects.get(
                        zuns_for_pr__work_program_change_in_discipline_block_module=practice_cb)
                except Practice.DoesNotExist:
                    continue
                credit_units_counter += sum([int(ze) for ze in practice_obj.ze_v_sem.split(", ")])
                if credit_units_counter <= int(module.selection_parametr):
                    process_practice(ws, practice_cb)
                else:
                    break
        if module.selection_rule == "all":
            for practice_cb in cbs:
                process_practice(ws, practice_cb)

    return wb_obj


"""def find_ze_modules_prac(modules=DisciplineBlockModule.objects.filter(descipline_block__name="Блок 2. Практика")):
    for module in modules:
        if module.selection_rule == "by_credit_units":
            print(module.id, module.name, module.selection_parametr)
        find_ze_modules_prac(module.childs.all())
"""
